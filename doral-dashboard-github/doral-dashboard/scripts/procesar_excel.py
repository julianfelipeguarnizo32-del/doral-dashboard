#!/usr/bin/env python3
"""
DORAL STORE — Procesador Automático de KPIs
Convierte el Excel de órdenes en data.json para el dashboard.
Se ejecuta automáticamente via GitHub Actions cuando se sube un Excel.
"""

import json, sys, os, glob
from datetime import datetime, date
from collections import defaultdict, Counter

# ── Intentar importar openpyxl ──
try:
    import openpyxl
except ImportError:
    os.system("pip install openpyxl --quiet")
    import openpyxl

# ══════════════════════════════════════════════
# 1. ENCONTRAR EL EXCEL MÁS RECIENTE
# ══════════════════════════════════════════════
def encontrar_excel():
    """Busca el Excel más reciente en la carpeta data/"""
    patrones = ["data/*.xlsx", "data/*.xls", "*.xlsx"]
    archivos = []
    for p in patrones:
        archivos.extend(glob.glob(p))
    if not archivos:
        sys.exit("❌ No se encontró ningún archivo Excel en data/")
    archivo = sorted(archivos, key=os.path.getmtime, reverse=True)[0]
    print(f"📂 Procesando: {archivo}")
    return archivo

# ══════════════════════════════════════════════
# 2. CARGAR DATOS
# ══════════════════════════════════════════════
def cargar_datos(archivo):
    wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    h = {v: i + 1 for i, v in enumerate(headers) if v}

    def pd(d):
        if not d: return None
        if isinstance(d, (datetime, date)):
            return d if isinstance(d, date) else d.date()
        for fmt in ['%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d']:
            try: return datetime.strptime(str(d).strip(), fmt).date()
            except: pass
        return None

    rows = []
    for r in range(2, ws.max_row + 1):
        row = {k: ws.cell(r, h[k]).value for k in h}
        rows.append(row)
    wb.close()

    orders = {}
    for row in rows:
        oid = row.get('ID')
        if oid and oid not in orders:
            orders[oid] = row

    # Detectar fecha de corte automáticamente
    fechas = [pd(o.get('FECHA')) for o in orders.values() if pd(o.get('FECHA'))]
    corte = max(fechas) if fechas else date.today()
    print(f"📅 Corte detectado: {corte}")

    return rows, orders, pd, corte

# ══════════════════════════════════════════════
# 3. CALCULAR TODOS LOS KPIs
# ══════════════════════════════════════════════
def calcular_kpis(rows, orders, pd, corte):
    nov_rows = [r for r in rows if r.get('NOVEDAD')]
    today = corte

    # ── GLOBAL ──
    total_ped = len(orders)
    total_fac = sum(o.get('TOTAL DE LA ORDEN') or 0 for o in orders.values())
    total_flete = sum(o.get('PRECIO FLETE') or 0 for o in orders.values())
    total_ent = sum(1 for o in orders.values() if o.get('ESTATUS') == 'ENTREGADO')
    total_dev = sum(1 for o in orders.values() if 'DEVOLUCION' in str(o.get('ESTATUS', '')))
    total_canc = sum(1 for o in orders.values() if o.get('ESTATUS') == 'CANCELADO')
    total_pend = sum(1 for o in orders.values() if o.get('ESTATUS') == 'PENDIENTE CONFIRMACION')
    total_desp = sum(1 for o in orders.values() if o.get('ESTATUS') not in {'CANCELADO', 'PENDIENTE CONFIRMACION'})
    dias_op = (corte - date(corte.year, 1, 1)).days + 1

    # ── POR MES ──
    ped_mes = {}; fac_mes = {}; flete_mes = {}; desp_mes = {}; ent_mes = {}
    dev_mes = {}; canc_mes = {}; pend_mes = {}

    for m in [1, 2, 3]:
        mo = [o for o in orders.values() if pd(o.get('FECHA')) and pd(o.get('FECHA')).month == m]
        if not mo:
            ped_mes[m] = fac_mes[m] = flete_mes[m] = desp_mes[m] = ent_mes[m] = 0
            dev_mes[m] = canc_mes[m] = pend_mes[m] = 0
            continue
        ped_mes[m] = len(mo)
        fac_mes[m] = round(sum(o.get('TOTAL DE LA ORDEN') or 0 for o in mo))
        flete_mes[m] = round(sum(o.get('PRECIO FLETE') or 0 for o in mo))
        desp_mes[m] = sum(1 for o in mo if o.get('ESTATUS') not in {'CANCELADO', 'PENDIENTE CONFIRMACION'})
        ent_mes[m] = sum(1 for o in mo if o.get('ESTATUS') == 'ENTREGADO')
        dev_mes[m] = sum(1 for o in mo if 'DEVOLUCION' in str(o.get('ESTATUS', '')))
        canc_mes[m] = sum(1 for o in mo if o.get('ESTATUS') == 'CANCELADO')
        pend_mes[m] = sum(1 for o in mo if o.get('ESTATUS') == 'PENDIENTE CONFIRMACION')

    # ── RENTABILIDAD ──
    rent = {}
    for m in [1, 2, 3]:
        mo = [o for o in orders.values() if pd(o.get('FECHA')) and pd(o.get('FECHA')).month == m]
        dev_mo = [o for o in mo if 'DEVOLUCION' in str(o.get('ESTATUS', ''))]
        fac = fac_mes.get(m, 0)
        flete = flete_mes.get(m, 0)
        flete_dev = round(sum(o.get('PRECIO FLETE') or 0 for o in dev_mo))
        val_dev = round(sum(o.get('TOTAL DE LA ORDEN') or 0 for o in dev_mo))
        ent_mo = [o for o in mo if o.get('ESTATUS') == 'ENTREGADO']
        cxe = round(flete / len(ent_mo)) if ent_mo else 0
        rent[m] = {
            'fac': fac, 'flete': flete,
            'flete_dev': flete_dev, 'val_dev': val_dev,
            'neto': round(fac - flete - flete_dev - val_dev),
            'ganancia': round(fac - flete - val_dev),
            'cxe': cxe
        }

    # ── NOVEDADES ──
    nov_kpis = {}
    for m in [1, 2, 3]:
        mo_nov = [r for r in nov_rows if pd(r.get('FECHA DE NOVEDAD')) and pd(r.get('FECHA DE NOVEDAD')).month == m]
        sol = sum(1 for r in mo_nov if str(r.get('FUE SOLUCIONADA LA NOVEDAD') or '').upper() == 'SI')
        ent_n = sum(1 for r in mo_nov if 'ENTREGADO' in str(r.get('ESTATUS') or ''))
        nov_kpis[m] = {
            'total': len(mo_nov), 'sol': sol,
            'nosol': len(mo_nov) - sol,
            'sol_pct': round(sol / len(mo_nov) * 100, 1) if mo_nov else 0,
            'ent': ent_n,
            'ent_pct': round(ent_n / len(mo_nov) * 100, 1) if mo_nov else 0
        }
    total_nov = sum(nov_kpis[m]['total'] for m in [1, 2, 3])
    total_sol = sum(nov_kpis[m]['sol'] for m in [1, 2, 3])
    total_ent_nov = sum(nov_kpis[m]['ent'] for m in [1, 2, 3])

    # ── AGING ──
    EXCL = {'ENTREGADO', 'CANCELADO', 'PENDIENTE CONFIRMACION'}
    en_t = [o for o in orders.values()
            if o.get('ESTATUS') not in EXCL and 'DEVOLUCION' not in str(o.get('ESTATUS', ''))]
    aging = {'1-3': 0, '4-7': 0, '8-14': 0, '15-30': 0, '>30': 0}
    por_trans_aging = defaultdict(list)
    for o in en_t:
        fg = pd(o.get('FECHA GUIA GENERADA')); fd = pd(o.get('FECHA'))
        ref = fg or fd
        if not ref: continue
        dias = (today - ref).days
        t = str(o.get('TRANSPORTADORA') or '').strip().upper()
        if dias <= 3: aging['1-3'] += 1
        elif dias <= 7: aging['4-7'] += 1
        elif dias <= 14: aging['8-14'] += 1
        elif dias <= 30: aging['15-30'] += 1
        else: aging['>30'] += 1
        if t: por_trans_aging[t].append(dias)

    aging_trans = {}
    for t in ['ENVIA', 'INTERRAPIDISIMO', 'VELOCES', 'COORDINADORA', 'TCC']:
        d = por_trans_aging.get(t, [])
        aging_trans[t] = {
            'n': len(d),
            'avg': round(sum(d) / len(d), 1) if d else 0,
            'alto': sum(1 for x in d if x > 7)
        }

    # ── PRIMERA ENTREGA ──
    nov_ids = {str(r.get('ID')) for r in rows if r.get('NOVEDAD')}
    primera_global = sum(1 for o in orders.values()
                         if o.get('ESTATUS') == 'ENTREGADO' and str(o.get('ID')) not in nov_ids)
    primera_mes = {}
    for m in [1, 2, 3]:
        mo = [o for o in orders.values() if pd(o.get('FECHA')) and pd(o.get('FECHA')).month == m]
        ent_mo = [o for o in mo if o.get('ESTATUS') == 'ENTREGADO']
        primera = sum(1 for o in ent_mo if str(o.get('ID')) not in nov_ids)
        primera_mes[m] = {
            'ent': len(ent_mo), 'primera': primera,
            'pct': round(primera / len(ent_mo) * 100, 1) if ent_mo else 0,
            'cancel': canc_mes[m],
            'cancel_pct': round(canc_mes[m] / ped_mes[m] * 100, 1) if ped_mes.get(m, 0) > 0 else 0
        }

    # ── TG ──
    tg_mes = {}
    for m in [1, 2, 3]:
        mo = [o for o in orders.values() if pd(o.get('FECHA')) and pd(o.get('FECHA')).month == m]
        desp_mo = [o for o in mo if o.get('ESTATUS') not in {'CANCELADO', 'PENDIENTE CONFIRMACION'}]
        tg_vals = [(pd(o.get('FECHA GUIA GENERADA')) - pd(o.get('FECHA'))).days
                   for o in desp_mo
                   if pd(o.get('FECHA GUIA GENERADA')) and pd(o.get('FECHA'))
                   and pd(o.get('FECHA GUIA GENERADA')) >= pd(o.get('FECHA'))]
        tg_mes[m] = {'avg': round(sum(tg_vals) / len(tg_vals), 2) if tg_vals else 0, 'n': len(tg_vals)}

    # ── FAC_DIA ──
    fac_dia = {}
    for o in orders.values():
        fd = pd(o.get('FECHA'))
        if fd:
            dk = fd.strftime('%Y-%m-%d')
            fac_dia[dk] = fac_dia.get(dk, 0) + (o.get('TOTAL DE LA ORDEN') or 0)
    fac_dia = {k: round(v) for k, v in sorted(fac_dia.items())}

    # ── TRANSPORTADORAS ──
    trans_kpi = {}
    te_trans = {}
    for t in ['ENVIA', 'INTERRAPIDISIMO', 'VELOCES', 'JAMV-DRIVE', 'COORDINADORA', 'TCC']:
        trans_kpi[t] = {}
        te_trans[t] = {}
        for m in [1, 2, 3]:
            mo = [o for o in orders.values()
                  if pd(o.get('FECHA')) and pd(o.get('FECHA')).month == m
                  and str(o.get('TRANSPORTADORA') or '').strip().upper() == t]
            if not mo:
                trans_kpi[t][f'2026-{m:02d}'] = {'gen': 0, 'desp': 0, 'tasa': 0, 'tg': 0}
                te_trans[t][f'2026-{m:02d}'] = {'total': 0, 'ent': 0, 'dev': 0, 'pct_ent': 0, 'pct_dev': 0}
                continue
            desp_t = sum(1 for o in mo if o.get('ESTATUS') not in {'CANCELADO', 'PENDIENTE CONFIRMACION'})
            ent_t = sum(1 for o in mo if o.get('ESTATUS') == 'ENTREGADO')
            dev_t = sum(1 for o in mo if 'DEVOLUCION' in str(o.get('ESTATUS', '')))
            tg_t = [(pd(o.get('FECHA GUIA GENERADA')) - pd(o.get('FECHA'))).days
                    for o in mo if pd(o.get('FECHA GUIA GENERADA')) and pd(o.get('FECHA'))
                    and pd(o.get('FECHA GUIA GENERADA')) >= pd(o.get('FECHA'))
                    and o.get('ESTATUS') not in {'CANCELADO', 'PENDIENTE CONFIRMACION'}]
            tg_avg = round(sum(tg_t) / len(tg_t), 2) if tg_t else 0
            tasa = round(desp_t / len(mo) * 100, 1) if mo else 0
            trans_kpi[t][f'2026-{m:02d}'] = {'gen': len(mo), 'desp': desp_t, 'tasa': tasa, 'tg': tg_avg}
            te_trans[t][f'2026-{m:02d}'] = {
                'total': len(mo), 'ent': ent_t, 'dev': dev_t,
                'pct_ent': round(ent_t / len(mo) * 100, 1) if mo else 0,
                'pct_dev': round(dev_t / len(mo) * 100, 1) if mo else 0
            }
            # Totales acumulados
            if 'total' not in te_trans[t]:
                t_all = [o for o in orders.values() if str(o.get('TRANSPORTADORA') or '').strip().upper() == t]
                ent_all = sum(1 for o in t_all if o.get('ESTATUS') == 'ENTREGADO')
                dev_all = sum(1 for o in t_all if 'DEVOLUCION' in str(o.get('ESTATUS', '')))
                te_trans[t]['total'] = {
                    'total': len(t_all), 'ent': ent_all, 'dev': dev_all,
                    'pct_ent': round(ent_all / len(t_all) * 100, 1) if t_all else 0,
                    'pct_dev': round(dev_all / len(t_all) * 100, 1) if t_all else 0
                }

    # ── TIENDA ──
    tienda_data = defaultdict(lambda: defaultdict(lambda: {'ped': 0, 'uds': 0, 'fac': 0}))
    seen = set()
    for row in rows:
        oid = row.get('ID')
        t = str(row.get('TIENDA') or '').strip().upper()
        m_d = pd(row.get('FECHA'))
        if not t or not m_d or not oid: continue
        m = m_d.month
        if oid not in seen:
            tienda_data[t][m]['ped'] += 1
            tienda_data[t][m]['fac'] += row.get('TOTAL DE LA ORDEN') or 0
            seen.add(oid)
        tienda_data[t][m]['uds'] += row.get('CANTIDAD') or 0

    lucidbot_uds = sum(tienda_data.get('1234296', {}).get(m, {}).get('uds', 0) for m in [1, 2, 3])
    lucidbot_fac = round(sum(tienda_data.get('1234296', {}).get(m, {}).get('fac', 0) for m in [1, 2, 3]))
    shopify_uds = sum(
        sum(tienda_data.get(t, {}).get(m, {}).get('uds', 0) for m in [1, 2, 3])
        for t in ['SHOPIFY NEW 2026', 'DORAL SHOPIFY 2026', 'DORAL SHOPIFY 2026 TOKEN',
                  'IMPORTACIONES METATRON', 'METATRON 2 - CHATEA PRO']
    )
    shopify_fac = round(sum(
        sum(tienda_data.get(t, {}).get(m, {}).get('fac', 0) for m in [1, 2, 3])
        for t in ['SHOPIFY NEW 2026', 'DORAL SHOPIFY 2026', 'DORAL SHOPIFY 2026 TOKEN',
                  'IMPORTACIONES METATRON', 'METATRON 2 - CHATEA PRO']
    ))

    # ── MARZO DIARIO FORECAST ──
    mar_daily = {d: v for d, v in fac_dia.items() if d.startswith('2026-03')}
    n_days_mar = len(mar_daily)
    total_fac_mar = fac_mes.get(3, 0)
    avg_dia = round(total_fac_mar / n_days_mar / 1e6, 2) if n_days_mar else 0
    avg_ped = round(ped_mes.get(3, 0) / n_days_mar, 1) if n_days_mar else 0
    forecast_fac = round(avg_dia * 31)
    forecast_ped = round(avg_ped * 31)

    # ── CONF_DAILY ──
    conf_daily = {}
    for o in orders.values():
        fd = pd(o.get('FECHA'))
        if not fd or fd.month != 3: continue
        dk = fd.strftime('%Y-%m-%d')
        if dk not in conf_daily:
            conf_daily[dk] = {'total': 0, 'pend': 0, 'conf': 0}
        conf_daily[dk]['total'] += 1
        if o.get('ESTATUS') == 'PENDIENTE CONFIRMACION':
            conf_daily[dk]['pend'] += 1
        else:
            conf_daily[dk]['conf'] += 1

    # Añadir pct
    for dk, d in conf_daily.items():
        d['pct_pend'] = round(d['pend'] / d['total'] * 100, 1) if d['total'] else 0
        d['pct_conf'] = round(d['conf'] / d['total'] * 100, 1) if d['total'] else 0
        d['uds_conf'] = d['conf']

    # ── GEO ──
    deps = Counter(str(o.get('DEPARTAMENTO DESTINO') or '').strip().upper()
                   for o in orders.values() if o.get('DEPARTAMENTO DESTINO'))
    cities_global = Counter(str(o.get('CIUDAD DESTINO') or '').strip().upper()
                            for o in orders.values() if o.get('CIUDAD DESTINO'))
    mar_ords = [o for o in orders.values() if pd(o.get('FECHA')) and pd(o.get('FECHA')).month == 3]
    cities_mar = Counter(str(o.get('CIUDAD DESTINO') or '').strip().upper()
                         for o in mar_ords if o.get('CIUDAD DESTINO'))
    top_dep = deps.most_common(1)[0] if deps else ('N/A', 0)
    top_city = cities_global.most_common(1)[0] if cities_global else ('N/A', 0)
    top_city_mar = cities_mar.most_common(1)[0] if cities_mar else ('N/A', 0)

    # ── CLIENTES ──
    tels = [str(o.get('TELÉFONO') or '') for o in orders.values() if o.get('TELÉFONO')]
    tc = Counter(tels)
    unicos = len(tc)
    recurrentes = sum(1 for _, c in tc.items() if c > 1)
    ped_rec = sum(c for _, c in tc.items() if c > 1)

    # ── NOV_DIAS ──
    nov_dias = {}
    for r in nov_rows:
        fn = pd(r.get('FECHA DE NOVEDAD'))
        if not fn or fn.month not in [1, 2, 3]: continue
        dk = fn.strftime('%Y-%m-%d')
        if dk not in nov_dias:
            nov_dias[dk] = {'total': 0, 'sol': 0, 'act': 0}
        nov_dias[dk]['total'] += 1
        if str(r.get('FUE SOLUCIONADA LA NOVEDAD') or '').upper() == 'SI':
            nov_dias[dk]['sol'] += 1
        else:
            nov_dias[dk]['act'] += 1

    # ── EST_GLOBAL ──
    est_global = dict(Counter(o.get('ESTATUS') for o in orders.values()).most_common(20))

    # ── PPTO DIARIO ──
    ppto = 500_000_000
    ppto_dia = {1: round(ppto/31), 2: round(ppto/28), 3: round(ppto/31)}
    ppto_daily = {}
    for mes_n, days_in_month in [(1, 31), (2, 28), (3, 31)]:
        mes_fac_dia = {d: v for d, v in fac_dia.items() if d.startswith(f'2026-{mes_n:02d}')}
        labels = []; fac_v = []; ppto_v = []; fac_acum = []; ppto_acum = []; pct_v = []
        cum_fac = 0; cum_ppto = 0
        p_dia = round(ppto / days_in_month)
        for dk in sorted(mes_fac_dia.keys()):
            day = int(dk[-2:])
            v = mes_fac_dia[dk]
            cum_fac += v; cum_ppto += p_dia
            labels.append(f"{day:02d}")
            fac_v.append(round(v/1e6, 2))
            ppto_v.append(round(p_dia/1e6, 2))
            fac_acum.append(round(cum_fac/1e6, 2))
            ppto_acum.append(round(cum_ppto/1e6, 2))
            pct_v.append(round(v/p_dia*100, 1))
        ppto_daily[mes_n] = {
            'labels': labels, 'fac': fac_v, 'ppto': ppto_v,
            'fac_acum': fac_acum, 'ppto_acum': ppto_acum,
            'pct': pct_v, 'ppto_mes': 500, 'dias': days_in_month,
            'ppto_dia': round(p_dia/1e6, 3)
        }

    # ══════════════════════════════════════════════
    # ENSAMBLAR JSON FINAL
    # ══════════════════════════════════════════════
    data = {
        'meta': {
            'corte': corte.strftime('%Y-%m-%d'),
            'corte_label': corte.strftime('%d de Marzo %Y') if corte.month == 3 else corte.strftime('%d/%m/%Y'),
            'corte_badge': corte.strftime('%d/%m/%Y'),
            'dias_operacion': dias_op,
            'generado': datetime.now().isoformat()
        },
        'global': {
            'ped': total_ped, 'fac': round(total_fac), 'flete': round(total_flete),
            'desp': total_desp, 'desp_pct': round(total_desp/total_ped*100, 1) if total_ped else 0,
            'ent': total_ent, 'ent_pct': round(total_ent/total_ped*100, 1) if total_ped else 0,
            'dev': total_dev, 'canc': total_canc, 'pend': total_pend
        },
        'mes': {
            str(m): {
                'ped': ped_mes.get(m, 0), 'fac': fac_mes.get(m, 0),
                'flete': flete_mes.get(m, 0), 'desp': desp_mes.get(m, 0),
                'desp_pct': round(desp_mes.get(m, 0)/ped_mes.get(m, 1)*100, 1) if ped_mes.get(m) else 0,
                'ent': ent_mes.get(m, 0),
                'ent_pct': round(ent_mes.get(m, 0)/ped_mes.get(m, 1)*100, 1) if ped_mes.get(m) else 0,
                'dev': dev_mes.get(m, 0), 'canc': canc_mes.get(m, 0), 'pend': pend_mes.get(m, 0)
            } for m in [1, 2, 3]
        },
        'rentabilidad': {str(m): rent[m] for m in [1, 2, 3]},
        'novedades': {
            'por_mes': {str(m): nov_kpis[m] for m in [1, 2, 3]},
            'total': total_nov, 'sol': total_sol,
            'nosol': total_nov - total_sol,
            'sol_pct': round(total_sol/total_nov*100, 1) if total_nov else 0,
            'ent': total_ent_nov,
            'ent_pct': round(total_ent_nov/total_sol*100, 1) if total_sol else 0
        },
        'aging': {
            'buckets': aging, 'por_trans': aging_trans, 'total': len(en_t)
        },
        'primera_entrega': {
            'global': {
                'ent': total_ent, 'primera': primera_global,
                'pct': round(primera_global/total_ent*100, 1) if total_ent else 0,
                'post_nov': total_ent - primera_global
            },
            'por_mes': {str(m): primera_mes[m] for m in [1, 2, 3]}
        },
        'tg': {str(m): tg_mes[m] for m in [1, 2, 3]},
        'fac_dia': fac_dia,
        'forecast': {
            'n_dias_real': n_days_mar,
            'avg_dia_ped': avg_ped,
            'avg_dia_fac': avg_dia,
            'forecast_ped': forecast_ped,
            'forecast_fac': forecast_fac
        },
        'trans_kpi': trans_kpi,
        'te_trans': te_trans,
        'tienda': {
            'lucidbot': {'uds': lucidbot_uds, 'fac': lucidbot_fac},
            'shopify': {'uds': shopify_uds, 'fac': shopify_fac}
        },
        'tienda_mes': {
            str(m): {t: {'uds': tienda_data[t][m]['uds'], 'fac': round(tienda_data[t][m]['fac'])}
                     for t in tienda_data if tienda_data[t][m]['ped'] > 0}
            for m in [1, 2, 3]
        },
        'conf_daily': conf_daily,
        'nov_dias': nov_dias,
        'est_global': est_global,
        'geo': {
            'total_deps': len(deps),
            'total_cities': len(cities_global),
            'top_dep': {'nombre': top_dep[0], 'ped': top_dep[1]},
            'top_city': {'nombre': top_city[0], 'ped': top_city[1]},
            'top_city_mar': {'nombre': top_city_mar[0], 'ped': top_city_mar[1]}
        },
        'clientes': {
            'unicos': unicos, 'recurrentes': recurrentes,
            'rec_pct': round(recurrentes/unicos*100, 1) if unicos else 0,
            'ped_recurrentes': ped_rec,
            'ped_rec_pct': round(ped_rec/total_ped*100, 1) if total_ped else 0
        },
        'ppto_diario': ppto_daily,
        'ppto_mensual': 500_000_000
    }

    return data

# ══════════════════════════════════════════════
# 4. MAIN
# ══════════════════════════════════════════════
def main():
    archivo = encontrar_excel()
    rows, orders, pd_fn, corte = cargar_datos(archivo)
    print(f"✅ {len(orders)} pedidos cargados")

    data = calcular_kpis(rows, orders, pd_fn, corte)
    print(f"✅ KPIs calculados — corte: {data['meta']['corte']}")

    # Guardar JSON
    os.makedirs('data', exist_ok=True)
    output = 'data/kpis.json'
    with open(output, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)

    size_kb = os.path.getsize(output) / 1024
    print(f"✅ {output} generado ({size_kb:.1f} KB)")
    print(f"📊 Pedidos: {data['global']['ped']} | Fac: ${data['global']['fac']/1e6:.1f}M | Corte: {data['meta']['corte']}")

if __name__ == '__main__':
    main()
